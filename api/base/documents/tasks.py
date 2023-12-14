from __future__ import absolute_import, unicode_literals

import os
import zipfile
import datetime
from django.conf import settings
from firma.celery import app
from openpyxl import load_workbook
from celery.utils.log import get_task_logger
from azure.core.exceptions import ResourceExistsError

from .models import Files, ZipFile, SignTask
from utils import document
from utils.odoo_client import OdooClient
from utils.azure_services import connect_to_azure_storage

logger = get_task_logger(__name__)


@app.task(name="send_contract_sign_task")
def send_contract_sign_task(files_id):
    logger.info("Start sending odoo contracts")

    # get files from database
    files = Files.objects.filter(id=files_id).first()

    '''
    contract_path = os.path.join(
        settings.MEDIA_ROOT+'/docs/', files.contract_name)
    employees_data_path = os.path.join(
       settings.MEDIA_ROOT+'/docs/', files.employees_data_name)
    '''

    # Send contract to odoo
    document.create_pdfs(files.contract_template , files.employees_data)
    logger.info("Odoo contracts sent")
    pass


@app.task(name="send_zip_file_task")
def send_zip_file_task(zip_task_id):
    '''
    Create a folder with the date inside media
    Extract Zip file with all the pdf files
    Open xls file with the data
    Iterate:
      Get the employee data from xls file: name, email, file name
      Get employee ID from Odoo, if not found, create new employee
      Send pdf file to Odoo
      Update PDF file with sign fields
      Tell Odoo to send the document via email to sign

    :param zip_file: Zip file with all the pdf files
    :param xls_file: XLS file with the data
    :param company_sign: True if the company sign is required
    '''
    print("Entré ya")
    logger.info("Starting task to send pdf files")

    # Get the files from database
    # Update the upload path from this documents
    # TOFIX: Add docs path as a variable, i.e, media/docs
    zip_task = ZipFile.objects.filter(id=zip_task_id).first()
    zip_file = f'media/{zip_task.zip_file}'
    xls_file = f'media/{zip_task.xlsx_file}'
    company_sign = zip_task.signs_number


    # Get system date and create a folder inside media folder with the date
    folder_name = 'contracts_' + datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    blob_service_client = connect_to_azure_storage()
    print("Conectado a Azure Storage: ",blob_service_client)

    try:
        os.mkdir('media/docs/' + folder_name)
    except FileExistsError as fError:
        logger.info(fError)


    with zipfile.ZipFile(zip_file, 'r') as zObject:
        file_list = zObject.infolist()
        zObject.extractall(path = 'media/docs/' + folder_name)
    
    # Get the first file in the list to get the path
    generated_dir = file_list[0].filename.split('/')[0]
    print("Generated dir: " + generated_dir)

    for contract_file in file_list:
        file_path = 'media/docs/' + folder_name + '/' + contract_file.filename
        print("File path: " + file_path)

        # Enter if the file exists
        if os.path.exists(file_path) and contract_file.filename.endswith('.pdf'):
            print("Entra a subir archivo a Azure")
            try:
                blob_client = blob_service_client.get_blob_client(container=settings.AZURE_STORAGE_CONTAINER, blob=file_path)
                print(f"Blob client: {str(blob_client)}")
                with open(file=file_path, mode='rb') as data:
                    print(f"###### Open binary file {type(data)}")
                    blob_client.upload_blob(data=data)

                logger.info("File uploaded to Azure: " + contract_file.filename)
                print(f"Upload to Azure: {file_path}")
                message = f"Archivo subido a Azure: {contract_file.filename}"
            except ResourceExistsError as rError:
                logger.info(rError)
                print(f"Error subiendo archivo a Azure {contract_file.filename}")
                message = f"Error subiendo archivo a Azure: {contract_file.filename}"
        else:
            print("No se envió archivos")
            blob_service_client.close()
            print(f"No subió archivos a Azure")
            message = "No subió archivos a Azure"
            # return {"error": "No se envió archivos"}
    
    # Close connection to Azure
    blob_service_client.close()

    xls = load_workbook(filename=xls_file, read_only=True, keep_vba=True, data_only=True)

    # get first sheet
    data_sheet = xls.worksheets[0]
        
    # Get column names
    column_names = [c.value for c in data_sheet[1] if c.value is not None]

    # Connect to Odoo
    odoo = OdooClient()
    print("Inicia proceso de Odoo")
     
    # Iterate to get the data
    for row in data_sheet.iter_rows(min_row=2, values_only=True):
        # Get employee data from XLSX File
        employee_data = {column_names[i]: row[i] for i in range(len(column_names))}

        # Get employee ID from Odoo
        employee_email = employee_data['CORREO'].strip()

        # Sometimes there's two emails in the same field, separated by comma
        # Check with angie if this is the correct way to handle this
        if ',' in employee_email:
            employee_email = employee_email.split(',')[0]
        
        employee_email = employee_email.lower()
        print('Correo empleado: '+employee_email)

        employee_odoo_id = odoo.search_employee(employee_email)

        # TODO: Pasar correo a minúsculas para evitar duplicados en Odoo
        if employee_odoo_id is None:
            employee_name = employee_data['NOMBRES Y APELLIDOS'].strip()
            print ('Creando empleado: ' + employee_name)
            employee_odoo_id = odoo.create_employee(employee_name, employee_email)

        # Get company signer ID from Odoo
        # TOFIX: Get company email from Odoo and search by email, maybe get directly the ID
        print(f'Firma Companía?: {company_sign}')
        # directorejectivo@fundacionudea.co

        # TODO: Cambiar el correo de la firma de la Fundación UdeA por una variable que llegue desde el form del frontend
        company_id = odoo.search_employee(settings.COMPANY_SIGNER) if company_sign == 2 else None
        
        # Transform PDF to base64, get number of pages to add sign fields in the last page
        nombre_archivo = employee_data['NOMBRE_ARCHIVO']

        # Full path to the file
        full_contract_path = 'media/docs/' + folder_name + '/' +  generated_dir + '/'+ nombre_archivo + '.pdf'
        try:
            document_64, numpages = document.convert_pdf_to_base64(full_contract_path)
            # Upload PDF file
            pdf_id = odoo.upload_new_contract_sign(nombre_archivo, document_64)
        except FileNotFoundError as fError:
            # Continua al siguiente archivo
            logger.info(fError)
            continue
        
        sign_task.message = f"Sube a Odoo documento {nombre_archivo}"
        sign_task.save()

        # Update PDF with sign fields
        second_field = True if company_sign == 2 else False
        sign_id = odoo.update_contract_sign(template_id=pdf_id, numpage=numpages, second_field=second_field)
        
        # Send document to sign
        odoo.send_sign_contract(pdf_id, nombre_archivo, employee_odoo_id, company_id)
        
        sign_task.last_contract_sent = nombre_archivo
        sign_task.message = f"Archivo envíado: {nombre_archivo}"
        
        files_sent += 1

        sign_task.files_sent = files_sent
        sign_task.save()

    sign_task.status = sign_task.STATUS_SUCCESS
    sign_task.save()
    return {"message": "Archivos enviados"}
