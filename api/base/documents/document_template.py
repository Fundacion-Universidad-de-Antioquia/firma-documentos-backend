'''
Sign.item.type().item_type('signature')
'''

from subprocess import Popen
from docxtpl import DocxTemplate
from openpyxl import load_workbook


class DocumentTemplate:

    def __init__(self, contract_file, employee_data_file):
        # Abrir Template docx
        self.template = DocxTemplate(contract_file)

        # abrir XLSX
        self.wb_obj = load_workbook(filename=employee_data_file)

    def convert_to_pdf(input_docx):
        LIBRE_OFFICE = '/usr/bin/libreoffice'
        RM = '/usr/bin/rm'
        p = Popen([LIBRE_OFFICE, '--headless', '--convert-to',
                   'pdf', '--outdir', 'docs', input_docx])

        p.communicate()

        p = Popen([RM, '-rf', input_docx])
        p.communicate()

    def get_fields_xlxs(self, file_path):
        '''
        Read xlxs File and get the column names
        :param file_path str
        '''

        wb_obj = load_workbook(filename=file_path)
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column

        # iterate through the columns to get the name on each cell
        columns = [sheet_obj.cell(row=1, column=x)
                   for x in range(1, max_col+1)]

        '''
        for i in range(1, max_col+1):
            columns.append(sheet_obj.cell(row=1, column=i))
        '''
        return columns

    def create_contracts_pdf(self):

        keys = self.get_fields_xlxs()

        max_col = self.sheet_obj.max_column

        column_names = self.get_fields_xlxs

        file_list = []

        # iterate through the columns to get the name on each cell
        for i in range(2, self.wb_obj.max_row+1):
            employee_data = [self.wb_obj.cell(row=1, column=x)
                             for x in range(i, max_col+1)]

            context = dict(zip(column_names), employee_data)

            contract_name = f'docs/{employee_data[1]}_contrato_{employee_data[2]}'
            new_document_name = contract_name+'docx'

            self.template.render(context)
            self.template.save(new_document_name)
            self.convert_to_pdf(new_document_name)
            file_list.append(contract_name+'.pdf')

        # Return the amount of contracts created
        return file_list

    def convert_to_pdf(self, input_docx):
        LIBRE_OFFICE = '/usr/bin/libreoffice'
        RM = '/usr/bin/rm'

        p = Popen([LIBRE_OFFICE, '--headless', '--convert-to',
                   'pdf', '--outdir', 'docs', input_docx])

        p.communicate()

        p = Popen([RM, '-rf', input_docx])
        p.communicate()

    def send_new_contract_sign(self, name, document_64):
        '''
        Create new contract registry with the PDF generated
        get the PDF document in base64
        Could be async method
        :param str name: project_cc.pdf
        :param base64 document_64:
        :return: ID of the document
        '''
        attachment_vals = [{
            'res_model': 'sign.template',
            'name': name,
            'type': 'binary',
            'res_id': 4567,
            'mimetype': 'application/pdf',
            'store_fname': name,
            'datas': document_64
        }]

        doc_id = self.api.execute_kw(self.database, self.uid, self.api_key,
                                     "ir.attachment", "create", attachment_vals)

        # domain = [name, f'data:application/pdf;base64,{document_64}', False]
        # doc_id = self.api.execute_kw(self.database, self.uid, self.api_key,
        #                             'sign.template', 'create', [domain])
        '''
        doc_id = self.api.execute_kw(self.database, self.uid, self.api_key,
                                     "sign.template", "create",
                                     [{'name': name, 'datas': document_64}])
        '''

        return doc_id

    def update_contract_sign(self, doc_name, document_id):
        '''
        Actualiza el contrato previamente creado con el campo de firma
        model = sign.template
        Item Type: Signature
        :param str doc_name: proyecto_cc.pdf
        :param str document_id: xxxx
            document_id: positivo: sign.template id en base de datos (ya debe estar en la BD)
        :return: Booleano True si agregó la firma
        '''
        sign_items = {
            'template_id': document_id,
            '0':
                {
                    'type_id': 1,
                    'required': True,
                    'name': 'Firma',
                    'responsible_id': 3,
                    'page': '2',
                    'posX': 0.065,
                    'posY': 0.811,
                    'width': 0.319,
                    'height': 0.051},
                '-1':
                    {
                    'type_id': 1,
                    'required': True,
                    'option_ids': [],
                    'responsible_id': 3,
                    'page': '2',
                    'posX': 0.659,
                    'posY': 0.813,
                    'width': 0.2,
                    'height': 0.05
                },
                'name': doc_name
        }

        # Create sign ITEM
        domain = [("sign.template")]

        result_sign = self.api.execute_kw(self.database, self.uid, self.api_key, "sign.template",
                                          "update_from_pdfviewer",
                                          sign_items)

        return result_sign
