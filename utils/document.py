import os
import base64
import PyPDF2
from subprocess import Popen
from docxtpl import DocxTemplate
from django.conf import settings
from openpyxl import load_workbook

from .odoo_client import OdooClient
from api.base.employees.models import Employee


def convert_pdf_to_base64(pdf_document):
    '''
    Convert PDF file to base 64
    Open the document and get the file encoded in base 64
    Get the total page numbers
    return a tuple (enconded_file, numpages)
    '''
    with open(pdf_document, "rb") as pdf_file:
        encoded_file = base64.b64encode(pdf_file.read())
        total_pages = len(PyPDF2.PdfReader(pdf_file).pages)

    pdf_data = encoded_file.decode('ascii')
    return (pdf_data, total_pages)


def convert_to_pdf(contract_path):
    '''
    Create the PDF file (contract) from the docx file
    '''
    LIBRE_OFFICE = '/usr/bin/libreoffice'
    RM = '/bin/rm'
    p = Popen([LIBRE_OFFICE, '--headless', '--convert-to',
              'pdf', '--outdir',  settings.MEDIA_ROOT + '/docs', contract_path+'.docx'])

    print([LIBRE_OFFICE, 'convierte a: ', 'pdf', contract_path+'.docx'])

    pdf_document = os.path.join(
        settings.MEDIA_ROOT+'/docs/', contract_path+'.pdf')

    p.communicate()

    p = Popen([RM, '-rf', contract_path+'.docx'])
    p.communicate()

    return pdf_document


# Open CSV File
def create_pdfs(template_file, employees_file):
    '''
    Read the excel file
    Get the first sheet
    Get the column names
    Create template for contract generation

    Iterate
        create new docx with the appropiate fields
        create new contract pdf
        Send new pdf to Odoo
        Update PDF with sign fields
        Tell Odoo    to send the document via email to sign
    '''

    xls = load_workbook(filename=employees_file, read_only=True, keep_vba=True)
    # get first sheet
    data_sheet = xls.worksheets[0]

    # Get column names from first file, has to filter the None values because sometimes it get the empty columns
    column_names = [c.value for c in data_sheet[1] if c.value is not None]
    print(f'Column names: {column_names}')

    odoo = OdooClient()

    for row in data_sheet.iter_rows(min_row=2, values_only=True):

        # Create the dictionary with context vars to create the contract
        document_context = {column_names[i]: row[i]
                            for i in range(len(column_names))}

        # Get the id, name and email from employee
        '''
        new_employee = Employee.objects.create(
            document_id=document_context['CEDULA_TRABAJADOR'],
            name=document_context['NOMBRE_TRABAJADOR'],
            email=document_context['CORREO_TRABAJADOR'])
        '''

        # row 0 = nombre, row[1] = Cédula
        document_name = document_context['NOMBRE_ARCHIVO']
        print(f'Document name: {document_name}')

        document_path = os.path.join(
            settings.MEDIA_ROOT+'/docs/', document_name)

        template = DocxTemplate(template_file)
        print(f'Render con Datos: {document_context}')
        template.render(document_context)
        template.save(document_path+'.docx')
        print('GUARDADO')

        # Create PDF from docx template
        pdf_document = convert_to_pdf(document_path)

        # Convert to base64 PDF, get a tuple (encoded_pdf, numpages)
        pdf_document_64, numpages = convert_pdf_to_base64(pdf_document)

        # Send document to Odoo
        odoo_document_id = odoo.upload_new_contract_sign(
            document_name, pdf_document_64)
        odoo_contract_sign = odoo.update_contract_sign(
            template_id=odoo_document_id, numpage=numpages)
        odoo_send_contract = odoo.send_sign_contract(
            document_name=document_name+'.pdf',
            employee_email=document_context['CORREO_TRABAJADOR'],
            template_id=odoo_document_id, 
            company_email=document_context['CORREO_FUNDACION']
        )

        # Create new record of document_signed
